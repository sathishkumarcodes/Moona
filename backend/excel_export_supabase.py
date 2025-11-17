"""
Excel export using Supabase
"""
from fastapi import APIRouter, HTTPException, Depends, Response
from fastapi.responses import StreamingResponse
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from auth_supabase import get_current_user_dependency, UserData
from db_supabase import get_db_pool, execute_query
import logging

logger = logging.getLogger(__name__)

export_router = APIRouter(prefix="/export", tags=["export"])

@export_router.get("/holdings/excel")
async def export_holdings_excel(
    current_user: UserData = Depends(get_current_user_dependency)
):
    """Export holdings to Excel file"""
    try:
        pool = await get_db_pool()
        
        # Fetch user's holdings
        holdings = await execute_query(
            "SELECT * FROM holdings WHERE user_id = $1 ORDER BY created_at DESC",
            current_user.id
        )
        
        if not holdings:
            raise HTTPException(status_code=404, detail="No holdings found")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Holdings sheet
        holdings_sheet = wb.create_sheet("Holdings")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        currency_format = '#,##0.00'
        percentage_format = '0.00%'
        number_format = '#,##0.00'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Holdings sheet headers
        holdings_headers = [
            "Symbol", "Name", "Type", "Sector", "Shares/Units", "Avg Cost", 
            "Current Price", "Total Cost", "Current Value", "Gain/Loss ($)", 
            "Gain/Loss (%)", "Last Updated"
        ]
        
        # Write headers
        for col, header in enumerate(holdings_headers, 1):
            cell = holdings_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write holdings data
        total_cost = 0
        total_value = 0
        
        for row, holding in enumerate(holdings, 2):
            holdings_sheet.cell(row=row, column=1, value=holding['symbol'])
            holdings_sheet.cell(row=row, column=2, value=holding['name'])
            holdings_sheet.cell(row=row, column=3, value=holding['type'].replace('_', ' ').title())
            holdings_sheet.cell(row=row, column=4, value=holding.get('sector') or "Other")
            
            shares_cell = holdings_sheet.cell(row=row, column=5, value=float(holding['shares']))
            shares_cell.number_format = number_format
            
            avg_cost_cell = holdings_sheet.cell(row=row, column=6, value=float(holding['avg_cost']))
            avg_cost_cell.number_format = currency_format
            
            current_price_cell = holdings_sheet.cell(row=row, column=7, value=float(holding['current_price']))
            current_price_cell.number_format = currency_format
            
            total_cost_cell = holdings_sheet.cell(row=row, column=8, value=float(holding['total_cost']))
            total_cost_cell.number_format = currency_format
            
            total_value_cell = holdings_sheet.cell(row=row, column=9, value=float(holding['total_value']))
            total_value_cell.number_format = currency_format
            
            gain_loss_cell = holdings_sheet.cell(row=row, column=10, value=float(holding['gain_loss']))
            gain_loss_cell.number_format = currency_format
            
            gain_loss_pct_cell = holdings_sheet.cell(row=row, column=11, value=float(holding['gain_loss_percent']) / 100)
            gain_loss_pct_cell.number_format = percentage_format
            
            # Color code gain/loss
            if holding['gain_loss'] >= 0:
                gain_loss_cell.font = Font(color="00B050")
                gain_loss_pct_cell.font = Font(color="00B050")
            else:
                gain_loss_cell.font = Font(color="C00000")
                gain_loss_pct_cell.font = Font(color="C00000")
            
            last_updated = holding.get('last_updated')
            if last_updated:
                if isinstance(last_updated, str):
                    last_updated_str = last_updated[:19]
                else:
                    last_updated_str = last_updated.strftime('%Y-%m-%d %H:%M:%S')
            else:
                last_updated_str = "N/A"
            
            holdings_sheet.cell(row=row, column=12, value=last_updated_str)
            
            # Apply borders
            for col in range(1, 13):
                holdings_sheet.cell(row=row, column=col).border = thin_border
            
            total_cost += float(holding['total_cost'])
            total_value += float(holding['total_value'])
        
        # Add summary row
        summary_row = len(holdings) + 3
        holdings_sheet.cell(row=summary_row, column=7, value="TOTALS:").font = Font(bold=True)
        holdings_sheet.cell(row=summary_row, column=8, value=total_cost).number_format = currency_format
        holdings_sheet.cell(row=summary_row, column=8).font = Font(bold=True)
        holdings_sheet.cell(row=summary_row, column=9, value=total_value).number_format = currency_format
        holdings_sheet.cell(row=summary_row, column=9).font = Font(bold=True)
        
        total_gain_loss = total_value - total_cost
        total_gain_loss_pct = (total_gain_loss / total_cost) * 100 if total_cost > 0 else 0
        
        gain_loss_total_cell = holdings_sheet.cell(row=summary_row, column=10, value=total_gain_loss)
        gain_loss_total_cell.number_format = currency_format
        gain_loss_total_cell.font = Font(bold=True, color="00B050" if total_gain_loss >= 0 else "C00000")
        
        gain_loss_pct_total_cell = holdings_sheet.cell(row=summary_row, column=11, value=total_gain_loss_pct / 100)
        gain_loss_pct_total_cell.number_format = percentage_format
        gain_loss_pct_total_cell.font = Font(bold=True, color="00B050" if total_gain_loss >= 0 else "C00000")
        
        # Auto-adjust column widths
        for col in range(1, 13):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in holdings_sheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            holdings_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create Summary sheet
        summary_sheet = wb.create_sheet("Portfolio Summary")
        
        summary_data = [
            ["Portfolio Summary", ""],
            ["", ""],
            ["Total Portfolio Value", total_value],
            ["Total Cost Basis", total_cost],
            ["Total Gain/Loss", total_gain_loss],
            ["Total Return %", total_gain_loss_pct / 100],
            ["", ""],
            ["Number of Holdings", len(holdings)],
            ["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row, column=1, value=label)
            if row == 1:
                summary_sheet.cell(row=row, column=1).font = Font(bold=True, size=16)
            elif label and value != "":
                summary_sheet.cell(row=row, column=1).font = Font(bold=True)
                summary_sheet.cell(row=row, column=2, value=value)
                
                if "Value" in label or "Cost" in label or "Gain/Loss" in label:
                    if label != "Total Return %":
                        summary_sheet.cell(row=row, column=2).number_format = currency_format
                    else:
                        summary_sheet.cell(row=row, column=2).number_format = percentage_format
                    
                    if "Gain/Loss" in label:
                        color = "00B050" if value >= 0 else "C00000"
                        summary_sheet.cell(row=row, column=2).font = Font(color=color, bold=True)
        
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 20
        
        # Asset allocation analysis
        asset_breakdown = {}
        for holding in holdings:
            asset_type = holding['type'].replace('_', ' ').title()
            if asset_type not in asset_breakdown:
                asset_breakdown[asset_type] = {"count": 0, "value": 0}
            asset_breakdown[asset_type]["count"] += 1
            asset_breakdown[asset_type]["value"] += float(holding['total_value'])
        
        start_row = len(summary_data) + 3
        summary_sheet.cell(row=start_row, column=1, value="Asset Breakdown").font = Font(bold=True, size=14)
        
        for i, (asset_type, data) in enumerate(asset_breakdown.items(), 1):
            row = start_row + i + 1
            summary_sheet.cell(row=row, column=1, value=f"{asset_type}:").font = Font(bold=True)
            
            percentage = (data["value"] / total_value) * 100 if total_value > 0 else 0
            summary_sheet.cell(row=row, column=2, value=f"{data['count']} holdings")
            summary_sheet.cell(row=row, column=3, value=data["value"]).number_format = currency_format
            summary_sheet.cell(row=row, column=4, value=percentage / 100).number_format = percentage_format
        
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 12
        
        # Save to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"portfolio_holdings_{timestamp}.xlsx"
        
        response = StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
        return response
    except Exception as e:
        logger.error(f"Error exporting holdings to Excel: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to export holdings")

